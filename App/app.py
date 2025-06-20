import streamlit as st
from streamlit_drawable_canvas import st_canvas
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import pandas as pd

# Archivos
EXCEL_FILE = "registros.xlsx"
CODIGOS_FILE = "codigos.txt"

# 🛡️ Asegurar archivo Excel
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Fecha", "Entrada", "Salida", "Firma", "Comentarios"])
    wb.save(EXCEL_FILE)

# 🎯 Cargar y guardar códigos
def cargar_codigos():
    if not os.path.exists(CODIGOS_FILE):
        return {}
    try:
        with open(CODIGOS_FILE, "r", encoding="utf-8") as f:
            return dict(line.strip().split(":", 1) for line in f if ":" in line)
    except UnicodeDecodeError:
        with open(CODIGOS_FILE, "r", encoding="latin1") as f:
            return dict(line.strip().split(":", 1) for line in f if ":" in line)


def guardar_codigo(codigo, nombre):
    with open(CODIGOS_FILE, "r", encoding="utf-8") as f:
        f.write(f"{codigo}:{nombre}\n")

# 🌸 Personalización visual
st.set_page_config(page_title="Kenia", page_icon="🌷", layout="centered")
st.markdown(
    """
    <style>
        body {
            background-color: #fff0f5;
        }
        .css-18e3th9 {
            background-color: #fff0f5 !important;
        }
        .stButton > button {
            background-color: #f78da7;
            color: white;
            border-radius: 10px;
            height: 3em;
            font-weight: bold;
        }
        .stRadio > div {
            background-color: #fce4ec;
            padding: 0.5em;
            border-radius: 8px;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# 🏢 Encabezado
st.markdown("## 🌟 **Unilever**")
st.markdown("#### 📋 Registro de Entrada y Salida de Personal")

# 📌 Cargar códigos
codigos = cargar_codigos()

# 📥 Entradas
codigo = st.text_input("🔐 Código (máx 8 caracteres)", max_chars=8)
nombre = ""
area = ""
nuevo = codigo not in codigos and codigo.strip() != ""

if codigo in codigos:
    nombre = codigos[codigo]
    st.success(f"👤 Bienvenida {nombre}")
else:
    nombre = st.text_input("🆕 Nombre (nuevo)", key="nombre")
    area = st.text_input("🏢 Área (nuevo)", key="area")

tipo = st.radio("🕓 Tipo de registro:", ["Entrada", "Salida"], horizontal=True)
comentarios = st.text_area("📝 Comentarios (opcional)")

# ✍️ Firma
st.markdown("### ✍️ Firma del empleado (obligatoria)")
canvas_result = st_canvas(
    fill_color="black",
    stroke_width=2,
    stroke_color="black",
    background_color="#ffffff",
    height=200,
    width=400,
    drawing_mode="freedraw",
    key="canvas"
)

firmado = (
    canvas_result.image_data is not None and
    canvas_result.json_data is not None and
    len(canvas_result.json_data["objects"]) > 0
)

# ✅ Registrar
if st.button("💾 Registrar"):
    if not codigo or len(codigo.strip()) > 8:
        st.error("❌ Código obligatorio (máx 8 caracteres)")
    elif nuevo and (not nombre or not area):
        st.error("❌ Nombre y área requeridos para nuevos registros")
    elif not firmado:
        st.error("❌ La firma es obligatoria para registrar")
    else:
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entrada = fecha if tipo == "Entrada" else ""
        salida = fecha if tipo == "Salida" else ""
        firma = "Sí"

        if nuevo:
            guardar_codigo(codigo, nombre)
            codigos[codigo] = nombre

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([nombre, fecha, entrada, salida, firma, comentarios])
        wb.save(EXCEL_FILE)

        st.success(f"✅ Registro exitoso para {nombre}")

        if canvas_result.image_data is not None:
            st.image(canvas_result.image_data, caption="🖋 Firma registrada", use_column_width=True)

# 🧾 Registros guardados
st.markdown("---")
st.subheader("📊 Registros guardados")

try:
    df = pd.read_excel(EXCEL_FILE)
    st.dataframe(df, use_container_width=True)
    with open(EXCEL_FILE, "rb") as file:
        st.download_button(
            label="📥 Descargar Excel",
            data=file,
            file_name=EXCEL_FILE,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
except Exception as e:
    st.error(f"❌ No se pudo cargar el archivo Excel: {e}")
