from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.textfield import MDTextField
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.label import MDLabel
from kivymd.uix.dialog import MDDialog
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.uix.card import MDCard
from kivy.uix.widget import Widget
from kivy.graphics import Color, Line
from kivy.core.window import Window
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "registros.xlsx"
CODIGOS_FILE = "codigos.txt"  # Para guardar código:nombre

# Asegurar que exista el archivo Excel
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Fecha", "Entrada", "Salida", "Firma", "Comentarios"])
    wb.save(EXCEL_FILE)

# Cargar códigos registrados
def cargar_codigos():
    if not os.path.exists(CODIGOS_FILE):
        return {}
    with open(CODIGOS_FILE, "r") as f:
        return dict(line.strip().split(":", 1) for line in f if ":" in line)

def guardar_codigo(codigo, nombre):
    with open(CODIGOS_FILE, "a") as f:
        f.write(f"{codigo}:{nombre}\n")

class FirmaWidget(Widget):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.firmado = False
        with self.canvas:
            Color(0, 0, 0)
        self.line = None

    def on_touch_down(self, touch):
        self.firmado = True
        with self.canvas:
            Color(0, 0, 0)
            self.line = Line(points=(touch.x, touch.y), width=2)

    def on_touch_move(self, touch):
        if self.line:
            self.line.points += [touch.x, touch.y]

    def limpiar(self):
        self.canvas.clear()
        self.firmado = False

class RegistroScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dialog = None
        self.codigos = cargar_codigos()

        layout = MDBoxLayout(orientation="vertical", padding=20, spacing=15)

        self.codigo = MDTextField(hint_text="Código (máx 8 caracteres)", mode="rectangle")
        self.nombre = MDTextField(hint_text="Nombre", mode="rectangle")
        self.area = MDTextField(hint_text="Área (solo nuevo)", mode="rectangle")
        self.comentarios = MDTextField(hint_text="Comentarios (opcional)", mode="rectangle")

        tipo_label = MDLabel(text="Tipo de registro:", theme_text_color="Secondary", halign="left")
        tipo_layout = MDBoxLayout(orientation="horizontal", spacing=30, size_hint_y=None, height=40)
        self.entrada = MDCheckbox(group="tipo", active=True)
        self.salida = MDCheckbox(group="tipo", active=False)
        tipo_layout.add_widget(MDLabel(text="Entrada")); tipo_layout.add_widget(self.entrada)
        tipo_layout.add_widget(MDLabel(text="Salida")); tipo_layout.add_widget(self.salida)

        self.firma = FirmaWidget(size_hint=(1, .3))

        self.boton_registrar = MDRaisedButton(text="Registrar", pos_hint={"center_x": .5}, on_release=self.registrar)

        layout.add_widget(self.codigo)
        layout.add_widget(self.nombre)
        layout.add_widget(self.area)
        layout.add_widget(tipo_label)
        layout.add_widget(tipo_layout)
        layout.add_widget(MDLabel(text="Firma del empleado (obligatoria):", halign="left"))
        layout.add_widget(self.firma)
        layout.add_widget(self.comentarios)
        layout.add_widget(self.boton_registrar)

        self.add_widget(layout)
        self.codigo.bind(text=self.verificar_codigo)

    def verificar_codigo(self, instance, value):
        value = value.strip()
        if value in self.codigos:
            self.nombre.text = self.codigos[value]
            self.area.disabled = True
        else:
            self.nombre.text = ""
            self.area.disabled = False

    def registrar(self, *args):
        codigo = self.codigo.text.strip()
        nombre = self.nombre.text.strip()
        area = self.area.text.strip()
        comentarios = self.comentarios.text.strip()
        tipo = "Entrada" if self.entrada.active else "Salida"
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not codigo or len(codigo) > 8:
            self.mostrar_dialogo("Código obligatorio (máx 8 caracteres)")
            return

        if codigo not in self.codigos and (not nombre or not area):
            self.mostrar_dialogo("Nombre y área requeridos para nuevos registros")
            return

        if not self.firma.firmado:
            self.mostrar_dialogo("La firma es obligatoria para registrar")
            return

        if codigo not in self.codigos:
            guardar_codigo(codigo, nombre)
            self.codigos[codigo] = nombre

        entrada = fecha if tipo == "Entrada" else ""
        salida = fecha if tipo == "Salida" else ""
        firmado = "Sí"

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([nombre, fecha, entrada, salida, firmado, comentarios])
        wb.save(EXCEL_FILE)

        self.mostrar_dialogo(f"Registro exitoso para {nombre}", limpiar=True)

    def mostrar_dialogo(self, mensaje, limpiar=False):
        if self.dialog:
            self.dialog.dismiss()
        self.dialog = MDDialog(
            title="Información",
            text=mensaje,
            buttons=[
                MDRaisedButton(text="Cerrar", on_release=lambda x: self.cerrar_dialogo(limpiar))
            ]
        )
        self.dialog.open()

    def cerrar_dialogo(self, limpiar):
        self.dialog.dismiss()
        if limpiar:
            self.codigo.text = ""
            self.nombre.text = ""
            self.area.text = ""
            self.comentarios.text = ""
            self.entrada.active = True
            self.salida.active = False
            self.firma.limpiar()

class RegistroApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Indigo"
        self.theme_cls.theme_style = "Light"
        return RegistroScreen()

if __name__ == "__main__":
    RegistroApp().run()
